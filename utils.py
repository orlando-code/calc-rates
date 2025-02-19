import yaml
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import unicodedata


# universal constants
MOLAR_MASS_CACO3 = 100.0869    # g/mol
PREFIXES = {'m': 1e-3, 'μ': 1e-6, 'n': 1e-9}
DURATIONS = {'hr': 24, 'd': 1, 'wk': 1/7}


# function to read in yamls
def read_yaml(yaml_fp) -> dict:
    with open(yaml_fp, "r") as stream:
        try:
            return yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)
            return None


def write_yaml(data: dict, fp="unnamed.yaml") -> None:
    with open(fp, "w") as file:
        yaml.dump(data, file)


def append_to_yaml(data: dict, fp="unnamed.yaml") -> None:
    with open(fp, "a") as file:
        yaml.dump(data, file)


def process_df(df: pd.DataFrame, require_results: bool=True, **selection_kws: dict) -> pd.DataFrame:
    # Default selection values
    default_selection = {'Extractor': 'Orlando', 'Include': 'yes'}
    
    # Merge defaults with user-provided values (user values take priority)
    selection_kws = {**default_selection, **selection_kws}
    
    # Apply selection
    for key, value in selection_kws.items():
        df = df[df[key] == value]
    
    # general processing
    df.columns = df.columns.str.lower() # columns lower case headers
    df.columns = df.columns.str.replace(' ', '_')   # process columns to replace whitespace with underscore
    df.columns = df.columns.str.replace('[()]', '', regex=True) # remove '(' and ')' from column names
    df['year'] = pd.to_datetime(df['year'], format='%Y')    # datetime format for later plotting
    df[['doi', 'year', 'authors']] = df[['doi', 'year', 'authors']].ffill()    # where I haven't these values for every row

    # df = df.map(lambda x: unicodedata.normalize("NFKC", str(x)).replace("μ", "u") if isinstance(x, str) else x)

    # missing values
    df = df[~df['n'].str.contains('~', na=False)]   # remove any rows in which 'n' has '~' in the string
    df = df[df.n != 'M']    # remove any rows in which 'n' is 'M'
    if require_results:
        df = df.dropna(subset=['n', 'calcification', 'calcification_units'])    # keep only rows with all the necessary data

    return df


def get_highlighted_mask(fp: str, sheet_name: str, rgb_color: str = "FFFFC000") -> pd.DataFrame:
    """
    Creates a boolean mask DataFrame where True indicates a highlighted cell.

    Args:
        fp (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).

    Returns:
        pd.DataFrame: Boolean mask DataFrame with True for highlighted cells.
    """
    wb = load_workbook(fp, data_only=True, read_only=True)
    ws = wb[sheet_name]

    df = pd.DataFrame([[cell.value for cell in row] for row in ws.iter_rows()])
    df = df.dropna(axis=1, how='all')   # remove any empty columns
    df.columns = df.iloc[0]  # set the first row as the column headers
    # df = df.drop(0)  # remove the header row
    mask = pd.DataFrame(False, index=df.index, columns=df.columns)  # mask with False by default, same shape as df
    # mask.drop(0, inplace=True)  # drop the first row (header)

    # set highlighted cells to True, all else False
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == rgb_color:
                mask.iat[row_idx, col_idx] = True

    return mask


def get_highlighted(fp: str, sheet_name: str, rgb_color: str = "FFFFC000", keep_cols: list = None) -> pd.DataFrame:
    """
    Loads an Excel sheet into a DataFrame and masks non-highlighted values with NaN,
    except for a specified subset of columns.

    Args:
        fp (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).
        keep_cols (list): List of column names to keep unchanged.

    Returns:
        pd.DataFrame: DataFrame with non-highlighted values masked as NaN.
    """
    # TODO: missing first two rows for some reason
    df = pd.read_excel(fp, sheet_name=sheet_name, engine="openpyxl", dtype=str)
    mask = get_highlighted_mask(fp, sheet_name, rgb_color)  # generate mask of highlighted cells

    if keep_cols:   # don't mask these columns
        keep_mask = pd.DataFrame(np.tile(df.columns.isin(keep_cols), (df.shape[0], 1)), 
                                index=df.index, columns=df.columns)
        final_mask = mask | keep_mask
    else:
        final_mask = mask

    return df.where(final_mask, np.nan)


def get_coordinates_from_gmaps(location_name: str, gmaps_client) -> pd.Series:
    """
    Get the latitude and longitude of a location using the Google Maps API. Used row-wise on a DataFrame.

    Args:
        location_name (str): Location name
        gmaps_client (googlemaps.client.Client): Google Maps API client. Requires an active, authorised Google Maps client.
    """
    try:
        result = gmaps_client.geocode(location_name)
        if result:
            lat = result[0]["geometry"]["location"]["lat"]
            lon = result[0]["geometry"]["location"]["lng"]
            return pd.Series([lat, lon])
        else:
            return pd.Series([None, None])
    except Exception as e:
        print(f"Error for {location_name}: {e}")
        return pd.Series([None, None])


def dms_to_dd(degrees, minutes=0, seconds=0, direction=""):
    """Convert degrees, minutes, and seconds to decimal degrees."""
    decimal_degrees = float(degrees) + float(minutes) / 60 + float(seconds) / 3600
    if direction in ["S", "W"]:
        decimal_degrees *= -1
    return decimal_degrees

def parse_coordinates(coord):
    """Convert various coordinate formats to decimal degrees."""
    coord = coord.strip()
    coord = coord.replace("''", '"')
    coord = coord.replace('"', "+")
    coord = coord.replace("′", "'")
    coord = coord.replace("’", "'")

    # **Decimal Degrees (DD)**
    dd_match = re.match(r"(-?\d+\.\d+)\s*[\u00B0]?\s*,?\s*(-?\d+\.\d+)\s*[\u00B0]?", coord)
    if dd_match:
        return float(dd_match.group(1)), float(dd_match.group(2))

    # **Degrees and Decimal Minutes (DMM)**
    dmm_match = re.match(
        r"(\d+\.?\d*)\u00B0\s*([NS]),?\s*(\d+\.?\d*)\u00B0\s*([EW])", coord
    )
    if dmm_match:
        lat_dd = dms_to_dd(dmm_match.group(1), direction=dmm_match.group(2))
        lon_dd = dms_to_dd(dmm_match.group(3), direction=dmm_match.group(4))
        return lat_dd, lon_dd

    # **Degrees, Minutes, and Seconds (DMS) & Degrees and Minutes (DM)**
    dms_match = re.match(
        r"(\d+)\u00B0\s*(\d+)[′']\s*([\d.]*)?[\″+]?\s*([NSEW])\D+"
        r"(\d+)\u00B0\s*(\d+)[′']\s*([\d.]*)?[\″+]?\s*([EWNS])", coord
    )
    if dms_match:
        lat_dd = dms_to_dd(dms_match.group(1), dms_match.group(2), dms_match.group(3) or 0, dms_match.group(4))
        lon_dd = dms_to_dd(dms_match.group(5), dms_match.group(6), dms_match.group(7) or 0, dms_match.group(8))
        return lat_dd, lon_dd

    raise ValueError(f"Unknown coordinate format: {coord}")


def rate_conversion(rate_val: float, rate_unit: str) -> float:
    """Conversion into gCaCO3 ... day-1 for absolute rates"""
    
    # convert moles to mass
    if 'mol' in rate_unit:
        rate_val *= MOLAR_MASS_CACO3
        mol_prefix = rate_unit.split('mol')[0]
        if mol_prefix in PREFIXES:
            rate_val *= PREFIXES[mol_prefix]

    # convert time unit
    for time_unit, factor in DURATIONS.items():
        if time_unit in rate_unit:
            rate_val *= factor
            break
    
    # area conversion
    if 'cm-2' in rate_unit.split(' ')[1]:
        rate_val *= 1e4   # cm2 to m2
    # mass conversion
    mass_prefix = rate_unit.split(' ')[1][0]
    if mass_prefix in PREFIXES:
        rate_val *= PREFIXES[mass_prefix]   # convert to g
    
    # TODO: add non-CaCO3 conversions

    
    return rate_val


def unit_name_conversion(rate_unit: str) -> str:
    # TODO: expand for all types of rate
    if "cm-2" in rate_unit:
        return "gCaCO3 m-2d-1"  # specific area unit
    elif "g-1" in rate_unit or "mg-1" in rate_unit:
        return "gCaCO3 g-1d-1"  # specific mass unit
    else:
        return "Unknown"
    
    
### sensitivity analysis
def select_by_stat(ds, variables_stats: dict):
    """
    Selects values from an xarray dataset based on the specified statistics for given variables.
    
    Parameters:
        ds (xr.Dataset): The input dataset with a 'param_combination' dimension and coordinates for the variables.
        variables_stats (dict): A dictionary where keys are variable names and values are the statistics to use for selection ('min', 'max', 'mean').
    
    Returns:
        xr.Dataset: Dataset subset at the selected coordinate values.
    """
    selected_coords = {}
    
    for var, stat in variables_stats.items():
        if stat == "min":
            selected_coords[var] = ds[var].min().item()
        elif stat == "max":
            selected_coords[var] = ds[var].max().item()
        elif stat == "mean":
            selected_coords[var] = ds[var].mean().item()
        else:
            raise ValueError(f"stat for {var} must be 'min', 'max', or 'mean'")
    
    # Select the closest matching values
    ds_selected = ds.sel(selected_coords, method="nearest")
    
    return ds_selected
