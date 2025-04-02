# files
import yaml
from openpyxl import load_workbook
import requests
from functools import lru_cache

# general
import numpy as np
import pandas as pd
import unicodedata
import re
import itertools
import string

# custom
from calcification import utils, config, locations
import cbsyst as cb
from tqdm.auto import tqdm
import cbsyst.helpers as cbh

### global constants
MOLAR_MASS_CACO3 = 100.0869    # g/mol
PREFIXES = {'c': 1e-2, 'm': 1e-3, 'μ': 1e-6, 'u': 1e-6, 'n': 1e-9} # TODO: probably shouldn't need both 'μ' and 'u'
DURATIONS = {'s': 86400, 'hr': 24, 'd': 1, 'wk': 1/7, 'month': 365/12, 'y': 1/365}


### file handling
# function to read in yamls
def read_yaml(yaml_fp) -> dict:
    with open(yaml_fp, "r") as stream:
        try:
            return yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)
            return None


def write_yaml(data: dict, fp="unnamed.yaml") -> None:
    # Convert numpy values to Python native types before serialization
    def _convert_numpy(obj):
        if isinstance(obj, dict):
            return {k: _convert_numpy(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [_convert_numpy(i) for i in obj]
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, np.number):
            return obj.item()
        elif isinstance(obj, (np.float32, np.float64)):
            return float(obj)
        elif isinstance(obj, (np.int32, np.int64)):
            return int(obj)
        else:
            return obj
            
    converted_data = _convert_numpy(data)
    with open(fp, "w") as file:
        yaml.dump(converted_data, file)


def append_to_yaml(data: dict, fp="unnamed.yaml") -> None:
    with open(fp, "a") as file:
        yaml.dump(data, file)


### processing files
def process_df(df: pd.DataFrame, require_results: bool=False, selection_dict: dict={'include': 'yes'}) -> pd.DataFrame:
    df.columns = df.columns.str.normalize("NFKC").str.replace("μ", "u") # replace any unicode versions of 'μ' with 'u'
    df = df.map(lambda x: unicodedata.normalize("NFKD", str(x)).replace('\xa0', ' ') if isinstance(x, str) else x)  # clean non-breaking spaces from string cells
    # general processing
    df.rename(columns=read_yaml(config.resources_dir / "mapping.yaml")['sheet_column_map'], inplace=True)    # rename columns to agree with cbsyst output    
    df.columns = df.columns.str.lower() # columns lower case headers for less confusing access later on
    df.columns = df.columns.str.replace(' ', '_')   # process columns to replace whitespace with underscore
    df.columns = df.columns.str.replace('[()]', '', regex=True) # remove '(' and ')' from column names
    df['year'] = pd.to_datetime(df['year'], format='%Y')    # datetime format for later plotting
    
    # flag up duplicate dois (only if they have also have 'include' as 'yes')
    inclusion_df = df[df['include'] == 'yes']
    duplicate_dois = inclusion_df[inclusion_df.duplicated(subset='doi', keep=False)]
    if not duplicate_dois.empty and not all(pd.isna(duplicate_dois['doi'])):
        print("\nDuplicate DOIs found, treat with caution:")
        print([doi for doi in duplicate_dois.doi.unique() if doi is not np.nan])
        
    # fill down necessary repeated metadata values
    df[['doi', 'year', 'authors', 'location', 'species_types', 'taxa']] = df[['doi', 'year', 'authors', 'location', 'species_types', 'taxa']].infer_objects(copy=False).ffill()
    df[['coords', 'cleaned_coords']] = df.groupby('doi')[['coords', 'cleaned_coords']].ffill()  # fill only as far as the next DOI
    
    if selection_dict:  # filter for selected values
        for key, value in selection_dict.items():
            df = df[df[key] == value]    
    
    df = uniquify_multilocation_study_dois(df)  # uniquify dois to reflect locations (for studies with multiple locations)
    # assign locations
    df = locations.assign_coordinates(df)  # assign coordinates to locations
    # save locations information
    locations.save_locations_information(df)
    
    # create family, genus, species, and functional group columns
    df = assign_taxonomical_info(df)
    
    # missing sample size values
    if df['n'].dtype == 'object':  # Only perform string operations if column contains strings
        df = df[~df['n'].str.contains('~', na=False)]   # remove any rows in which 'n' has '~' in the string
        df = df[df.n != 'M']    # remove any rows in which 'n' is 'M'
    df.loc[:, df.columns != 'year'] = df.loc[:, df.columns != 'year'].apply(safe_to_numeric)
    
    
    problem_cols = ['irr', 'ipar']  # some columns have rogue strings when they should all contain numbers: in this case, convert unconvertable values to NaN
    for col in problem_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df['irr'] = df.apply(lambda row: irradiance_conversion(row['ipar'], 'PAR') if pd.notna(row['ipar']) else row['irr'], axis=1)  # convert irradiance to consistent unit
    
    if require_results:
        df = df.dropna(subset=['n', 'calcification'])    # keep only rows with all the necessary data
    
    # calculate calcification standard deviation only when 'calcification_se' and 'n' are not NaN
    df['calcification_sd'] = df.apply(lambda row: calc_sd_from_se(row['calcification_se'], row['n']) if pd.notna(row['calcification_se']) and pd.notna(row['n']) else row['calcification_sd'], axis=1)
    
    # calculate standarised calcification rates and relevant units
    df = map_units(df)  # map units to st_units
    df[['st_calcification', 'st_calcification_sd', 'st_calcification_unit']] = df.apply(
        lambda x: pd.Series(utils.rate_conversion(x['calcification'], x['calcification_sd'], x['st_calcification_unit'])) if pd.notna(x['calcification']) and pd.notna(x['st_calcification_unit']) else pd.Series(['', '', '']), 
        axis=1)
    
    return df
        

def map_units(df: pd.DataFrame) -> pd.DataFrame:
    map_dict = utils.read_yaml(config.resources_dir / "mapping.yaml")["unit_map"]
    inverted_map = {val: key for key, values in map_dict.items() for val in values}

    df['st_calcification_unit'] = df['calcification_unit'].map(inverted_map)
    return df
    

def assign_taxonomical_info(df: pd.DataFrame) -> pd.DataFrame:
    df = df.assign(
        genus=df.species_types.apply(lambda x: binomial_to_genus_species(x)[0]),
        species=df.species_types.apply(lambda x: binomial_to_genus_species(x)[1])
    )   # separate binomials into genus and species columns
    # append family column: if no species_mapping file, generate
    if not (config.resources_dir / 'species_mapping.yaml').exists():
        create_species_mapping_yaml(df.species_types.unique())
    else:
        print(f'Using species mapping in {config.resources_dir / 'species_mapping.yaml'}.')
    species_mapping = read_yaml(config.resources_dir / 'species_mapping.yaml')
    # extract nested dictionary values for each species
    df['family'] = df.species_types.apply(lambda x: species_mapping.get(x, {}).get('family', 'Unknown'))
    df['functional_group'] = df.species_types.apply(lambda x: species_mapping.get(x, {}).get('functional_group', 'Unknown'))
    df['core_grouping'] = df.species_types.apply(lambda x: species_mapping.get(x, {}).get('core_grouping', 'Unknown'))
    return df


def uniquify_multilocation_study_dois(df: pd.DataFrame) -> pd.DataFrame:

    temp_df = df.copy()
    temp_df['original_doi'] = temp_df['doi']
    temp_df['location_lower'] = temp_df['location'].str.lower()

    # locs_df = temp_df.drop_duplicates(['doi', 'location_lower', 'cleaned_coords']).set_index('doi')
    locs_df = temp_df.drop_duplicates(['doi', 'location_lower', 'coords', 'cleaned_coords'])

    locs_df.loc[:,'doi'] = utils.uniquify_repeated_values(locs_df.doi)

    temp_df = temp_df.merge(locs_df['doi'], how='left', left_index=True, right_index=True, suffixes=("_old",""))
    # drop original doi column
    temp_df.drop(columns=['doi_old'], inplace=True)
    # group by original doi to fill down the new doi
    temp_df['doi'] = temp_df.groupby('original_doi')['doi'].ffill()
    return temp_df





def get_species_info_from_worms(species_binomial: str) -> dict:
    """Query WoRMS API to get the family name of a coral species.
    
    Args:
        species_binomial (str): Scientific name in 'Genus species' format
    
    Returns:
        dict: Dictionary with family name and additional taxonomic information
    """
    
    # strip leading/trailing whitespace
    species_binomial = species_binomial.strip()
    # clean species name
    if any(x in species_binomial.replace('.', '').split() for x in ['sp', 'spp', 'cf']): # remove general indication from species, leaving just genus
        species_binomial = species_binomial.split()[0]  # take the first word as the genus
    # manual filtering
    if species_binomial in ['Massive porites', 'Porites lutea/lobata']:
        species_binomial = 'Porites'
    if 'CCA' in species_binomial:   # e.g. 'Unknown CCA'
        return {'species': species_binomial, 'family': 'Corallinaceae-Sporolithaceae', 'status': 'Best guess', 'functional_group': 'Crustose coralline algae', 'core_grouping': 'CCA'}
    
    # URL encode the species name to handle spaces properly
    encoded_species = requests.utils.quote(species_binomial)
    base_url = f"https://www.marinespecies.org/rest/AphiaRecordsByName/{encoded_species}?like=false&marine_only=true"
    
    try:
        response = requests.get(base_url)
        response.raise_for_status()  # raise exception for 4XX/5XX status codes
        data = response.json()

        
        if data and isinstance(data, list) and len(data) > 0:   # assign response to dictionary if data found
            if len(data) > 1:   # if more than one record, take the most recent one which is 'accepted' in the 'status' field
                data = [record for record in data if record.get('status', '') == 'accepted'][0]
                if not data:    # if technically none accepted, take the first record (haven't seen this happen yet)
                    data = data[0]
            else:
                data = data[0]  # list containing single record dictionary: take the dictionary

            result = {
                'species': species_binomial,
                'genus': data.get('genus', 'Not Found'),
                'family': data.get('family', 'Not Found'),
                'status': 'Found',
                'rank': data.get('rank', 'Not Found'),
                'aphia_id': data.get('AphiaID', 'Not Found'),
                'accepted_name': data.get('valid_name', species_binomial),
                'kingdom': data.get('kingdom', 'Not Found'),
                'phylum': data.get('phylum', 'Not Found'),
                'class': data.get('class', 'Not Found'),
                'order': data.get('order', 'Not Found')
            }
            result['functional_group'] = assign_functional_group(result)
            result['core_grouping'] = assign_core_groupings(result)

            return result
        else:
            return {'species': species_binomial, 'family': 'Not Found', 'status': 'No Data', 'functional_group': 'Unknown', 'core_grouping': 'Unknown'}
    
    except requests.exceptions.RequestException as e:
        print(f"API Request Error for {species_binomial}: {e}")
        return {'species': species_binomial, 'family': 'Error', 'status': str(e), 'functional_group': 'Unknown', 'core_grouping': 'Unknown'}


def create_species_mapping_yaml(species_list) -> None:
    """Create a YAML file with species-family mapping for a list of species.
    
    Args:
        species_list (list): List of coral species names in 'Genus species' format
    """
    species_mapping = {}
    for species in tqdm(species_list, desc='Querying WoRMS API to retrieve organism taxonomic data'):
        species_info = get_species_info_from_worms(species)
        # print(species_info)
        species_mapping[species] = {
            'family': species_info['family'],
            'functional_group': species_info['functional_group'],
            'core_grouping': species_info['core_grouping']}
    # save family: genus, species mapping to YAML file
    with open(config.resources_dir / 'species_mapping.yaml', 'w') as file:
        yaml.dump(species_mapping, file)
    print(f'Species mapping saved to {config.resources_dir / "species_mapping.yaml"}')
    
    
def assign_functional_group(taxon_info):
    """
    Assign a functional group based on taxonomic information.
    N.B. this mapping is not exhaustive: only checked with ~130 species.
    There is also likely subjectivity in assignment.

    Args:
        taxon_info (dict): Dictionary containing taxonomic information
    
    Returns:
        str: Functional group name
    """
    family = taxon_info.get('family', '').lower()
    order = taxon_info.get('order', '').lower()
    class_name = taxon_info.get('class', '').lower()
    phylum = taxon_info.get('phylum', '').lower()
    genus = taxon_info.get('genus', '').lower()  # Extract genus from species name
    binomial = taxon_info.get('species', '').lower()

    if genus in ['jania', 'amphiroa']:
        return 'Articulate coralline algae'
    
    # Crustose coralline algae
    if family in ['corallinaceae', 'sporolithaceae', 'hapalidiaceae', 'hydrolithaceae', 'lithophyllaceae', 'mesophyllumaceae', 'spongitidaceae', 'porolithaceae']:
        return 'Crustose coralline algae'
    
    # Fleshy algae
    if (phylum in ['chlorophyta', 'ochrophyta', 'rhodophyta'] or 
        order in ['dictyotales', 'ectocarpales', 'fucales', 'gigartinales'] or
        class_name in ['phaeophyceae', 'ulvophyceae', 'florideophyceae']):
        # Special check for calcareous algae that aren't CCA
        if genus in ['galaxaura', 'padina']:
            return 'Calcareous algae'
        if genus in ['peyssonnelia']:
            return 'Calcareous red algae'
        if family in ['halimedaceae']:
            return 'Calcareous green algae'
        return 'Fleshy algae'
    
    # Hard corals (scleractinian)
    if order == 'scleractinia' or family in ['pocilloporidae', 'acroporidae', 'poritidae', 'faviidae', 'fungiidae', 'agariciidae']:
        return 'Hard coral'
    
    # Soft corals
    if order in ['alcyonacea', 'gorgonacea'] or 'alcyoniidae' in family:
        return 'Soft coral'
    
    # Sponges
    if phylum == 'porifera':
        return 'Sponge'
    
    # Foraminifera
    if phylum in ['foraminifera', 'retaria'] or class_name == 'foraminifera':
        return 'Foraminifera'
    
    # Turf algae - often identified by growth form rather than taxonomy
    if 'turf' in taxon_info.get('species', ''):
        return 'Turf algae'
    
    # Bryozoans
    if phylum == 'bryozoa':
        return 'Bryozoan'
    
    # catch-all case
    return 'Other benthic organism'
    
    
def assign_core_groupings(taxon_info: dict) -> str:
    """
    Assign a core grouping (CCA, halimeda, coral, foraminifera, other) based on taxonomical information.
    
    Args:
        taxon_info (dict): Functional group name
    
    Returns:
        str: Core grouping name
    """
    # check for halimeda in genus
    if taxon_info.get('genus', '').lower() == 'halimeda':
        return 'Halimeda'    
    
    if taxon_info['functional_group'] in ['Crustose coralline algae', 'Calcareous algae'] or 'calcareous' in taxon_info['functional_group'].lower():
        return 'CCA'        
    elif taxon_info['functional_group'] in ['Fleshy algae', 'Turf algae', 'Articulate coralline algae']:
        return 'Other algae'
    elif taxon_info['functional_group'] in ['Hard coral', 'Soft coral']:
        return 'Coral'
    elif taxon_info['functional_group'] in ['Foraminifera']:
        return 'Foraminifera'
    else:
        return 'Other'

def binomial_to_genus_species(binomial):
    """
    Convert a binomial name to genus and species.
    
    Args:
        binomial (str): Binomial name.
        
    Returns:
        tuple: Genus and species names.
    """
    # strip periods, 'cf' (used to compare with known species)
    binomial = binomial.replace('.', '')
    binomial = binomial.replace('cf', '')
    split = binomial.split(' ')
    # remove any empty strings (indicative of leading/trailing whitespace)
    split = [s for s in split if s]
    
    if 'spp' in binomial or 'sp' in split:
        genus = split[0]
        species = 'spp'
    else:
        genus = split[0]
        species = split[-1] if len(split) > 1 else 'spp'
    return genus, species


def aggregate_df(df, method: str='mean') -> pd.DataFrame:
    # Define aggregation functions
    aggregation_funcs = {col: method if pd.api.types.is_numeric_dtype(df[col]) else lambda x: x.iloc[0] for col in df.columns}

    # Aggregate DataFrame
    return df.agg(aggregation_funcs)


def calc_sd_from_se(se: float, n: int) -> float:
    """Calculate standard deviation from standard error and sample size
    
    Args:
        se (float): standard error
        n (int): number of samples
        
    Returns:
        float: standard deviation
    """
    return se * np.sqrt(n)


def safe_to_numeric(col):
    """Convert column to numeric if possible, otherwise return as is."""
    try:
        return pd.to_numeric(col)
    except (ValueError, TypeError):
        return col  # Return original column if conversion fails
    
    
def extract_year_from_str(s) -> pd.Timestamp:
    """Extract year from a string. Useful for cases where 'authors' has been provided with attached year.
    
    Args:
        s (str): String containing year.
        
    Returns:
        pd.Timestamp: Year extracted from string.
    """
    try:
        match = re.search(r'\d{4}', str(s))
        if match:
            return pd.to_datetime(match.group(), format='%Y')
        return pd.NaT
    except ValueError:
        return None


def irradiance_conversion(irr_val: float, irr_unit: str="PAR") -> float:
    # convert from mol quanta m-2 day-1 to μmol quanta m-2 s-1
    s_in_day = DURATIONS['s']
    return irr_val / (s_in_day * PREFIXES['μ']) if irr_unit == "PAR" else irr_val


def get_highlighted_mask(fp: str, sheet_name: str, rgb_color: str = "FFFFC000") -> pd.DataFrame:
    """
    Creates a boolean mask DataFrame where True indicates a highlighted cell.

    Args:
        fp (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).

    Returns:
        pd.DataFrame: Boolean mask DataFrame with True for highlighted cells.
    """
    wb = load_workbook(fp, data_only=True, read_only=True)
    ws = wb[sheet_name]

    df = pd.DataFrame([[cell.value for cell in row] for row in ws.iter_rows()])
    df = df.dropna(axis=1, how='all')   # remove any empty columns
    df.columns = df.iloc[0]  # set the first row as the column headers
    # df = df.drop(0)  # remove the header row
    mask = pd.DataFrame(False, index=df.index, columns=df.columns)  # mask with False by default, same shape as df
    # mask.drop(0, inplace=True)  # drop the first row (header)

    # set highlighted cells to True, all else False
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == rgb_color:
                mask.iat[row_idx, col_idx] = True
                
    # Always mark 'include' and 'n' columns as True for future processing
    for col in ['Include', 'n', 'Species types', 'Location']:
        if col in mask.columns:
            mask[col] = True

    return mask


def get_highlighted(fp: str, sheet_name: str, rgb_color: str = "FFFFC000") -> pd.DataFrame:
    """
    Loads an Excel sheet into a DataFrame and masks non-highlighted values with NaN,
    except for a specified subset of columns.

    Args:
        fp (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).
        keep_cols (list): List of column names to keep unchanged.

    Returns:
        pd.DataFrame: DataFrame with non-highlighted values masked as NaN.
    """
    df = pd.read_excel(fp, sheet_name=sheet_name, engine="openpyxl", dtype=str)
    mask = get_highlighted_mask(fp, sheet_name, rgb_color)  # generate mask of highlighted cells
    return df.where(mask.drop(0).reset_index(drop=True), np.nan)     # remove first row of mask and reset index


def uniquify_repeated_values(vals: list) -> list:
    """
    Append a unique suffix to repeated values in a list.
    
    Parameters:
        vals (list): List of values.
    
    Returns:
        list: List of values with unique suffixes.
    """
    def zip_letters(l):
        """Zip a list of strings with uppercase letters."""
        al = string.ascii_uppercase
        return ['-LOC-'.join(i) for i in zip(l, al)] if len(l) > 1 else l
    return [j for _, i in itertools.groupby(vals) for j in zip_letters(list(i))]


### carbonate chemistry
# @lru_cache(maxsize=32)
def populate_carbonate_chemistry(fp: str, sheet_name: str="all_data", selection_dict: dict={'include': 'yes'}) -> pd.DataFrame:
    df = process_df(pd.read_excel(fp, sheet_name=sheet_name), require_results=False, selection_dict=selection_dict)
    ### load measured values
    print("Loading measured values...")
    measured_df = get_highlighted(fp, sheet_name=sheet_name)    # keeping all cols
    # return measured_df
    measured_df = process_df(measured_df, require_results=False, selection_dict=selection_dict)
    
    ### convert nbs values to total scale using cbsyst     # TODO: implement uncertainty propagation
    print("Calculating total pH values...")
    # if phtot and phnbs are both nan, calculate phtot from temp, salinity, dic, and ta
    measured_df.loc[:, 'phtot'] = measured_df.apply(
        lambda row: cb.cbsyst.Csys(
            TA=row['ta'], DIC=row['dic'], T_in=row['temp'], S_in=row['sal'] if pd.notna(row['sal']) else 35,
        ).get('pHtot', None) if pd.isna(row['phnbs']) and pd.isna(row['phtot']) and pd.notna(row['temp']) and pd.notna(row['dic']) and pd.notna(row['ta'])
        else np.nan,
        axis=1
    )
    # if one of ph is provided, ensure total ph is calculated
    measured_df.loc[:, 'phtot'] = measured_df.apply(
        lambda row: cbh.pH_scale_converter(
            pH=row['phnbs'], scale='NBS', Temp=row['temp'], Sal=row['sal'] if pd.notna(row['sal']) else 35
        ).get('pHtot', None) if pd.notna(row['phnbs']) and pd.notna(row['temp'])
        else row['phtot'],
        axis=1
    )
    # count number of non-nan phtot values
    # print(f"\nTotal number of total pH values: {measured_df['phtot'].count()}")

    
    ### calculate carbonate chemistry
    carb_metadata = read_yaml(config.resources_dir / "mapping.yaml")
    carb_chem_cols = carb_metadata['carbonate_chemistry_cols']
    out_values = carb_metadata['carbonate_chemistry_params']
    carb_df = measured_df[carb_chem_cols].copy()

    # apply function row-wise
    tqdm.pandas(desc="Calculating carbonate chemistry")
    carb_df.loc[:, out_values] = carb_df.progress_apply(lambda row: pd.Series(calculate_carb_chem(row, out_values)), axis=1)
    return df.combine_first(carb_df)
    # return measured_df


def calculate_carb_chem(row, out_values: list) -> dict:
    """Calculate carbonate chemistry parameters and return a dictionary."""
    try:
        out_dict = cb.Csys(
            pHtot=row['phtot'],
            TA=row['ta'],
            T_in=row['temp'],
            S_in=row['sal'],
        )
        out_dict = {key.lower(): value for key, value in out_dict.items()}  # lower the keys of the dictionary to ensure case-insensitivity

        return {
            key: (out_dict.get(key.lower(), None)[0] if isinstance(out_dict.get(key.lower()), (list, np.ndarray)) else out_dict.get(key.lower(), None))
            for key in out_values
            }    
    except Exception as e:
        print(f"Error: {e}")


### UNIT STANDARDISATION
def parse_unit_components(unit: str) -> tuple[str, str]:
    """Parse a unit string into numerator and denominator components."""
    if ' ' not in unit:
        raise ValueError(f"Unit '{unit}' does not have proper format 'numerator denominator'")
    
    components = unit.split(' ')
    if len(components) != 2:
        raise ValueError(f"Unit '{unit}' has more than two components")
    
    return components[0], components[1]

def extract_prefix(unit_part: str, unit_type: str) -> tuple[str, str]:
    """Extract the prefix from a unit part (e.g., 'mg' -> 'm', 'g')."""
    if 'delta' in unit_part:    # relative changes
        return '', unit_part
    
    # determine unit type
    if unit_type == 'mass':
        match = re.match(r'([nuμmcdk]?)(g|mol)', unit_part)
    elif unit_type == 'area':
        match = re.match(r'([nuμmcdk]?)(m2|m-2)', unit_part)
    elif unit_type == 'length':
        match = re.match(r'([nuμmcdk]?)(m{1}$)', unit_part)
    else:
        return '', unit_part
    
    if match:
        return match.group(1), match.group(2)
    return '', unit_part


def convert_numerator(num_part: str, rate_val: float) -> tuple[float, str]:
    """Convert the numerator part of the unit."""
    if 'delta' in num_part: # relative changes
        return rate_val, num_part
    
    has_caco3 = 'CaCO3' in num_part    
    if has_caco3:
        num_part_clean = num_part.replace('CaCO3', '')
    else:
        num_part_clean = num_part
    
    if 'mol' in num_part_clean: # molar units
        prefix, base = extract_prefix(num_part_clean, 'mass')
        rate_val *= MOLAR_MASS_CACO3 * PREFIXES.get(prefix, 1.0)
        new_unit = 'g'
    elif 'g' in num_part_clean: # mass units
        prefix, base = extract_prefix(num_part_clean, 'mass')
        rate_val *= PREFIXES.get(prefix, 1.0)
        new_unit = 'g'
    elif 'm2' in num_part_clean:    # area units
        prefix, base = extract_prefix(num_part_clean, 'area')
        rate_val *= PREFIXES.get(prefix, 1.0) ** 2
        new_unit = 'm2'
    elif re.search(r'm{1}$', num_part_clean):   # extension units
        prefix, base = extract_prefix(num_part_clean, 'length')
        rate_val *= PREFIXES.get(prefix, 1.0)
        new_unit = 'm'
    elif re.match(r'[nuμmcdk]{2}', num_part_clean): # duplicate units e.g. mm
        prefix = num_part_clean[0]
        rate_val *= PREFIXES.get(prefix, 1.0)
        new_unit = num_part_clean[1:]
    else:
        new_unit = num_part_clean
    
    if has_caco3:   # TODO: extend to all?
        new_unit = f"{new_unit}CaCO3"
    
    return rate_val, new_unit

def convert_denominator(denom_part: str, rate_val: float) -> tuple[float, str]:
    """Convert the denominator part of the unit."""
    for duration, factor in DURATIONS.items():
        if duration in denom_part:
            rate_val *= factor
            denom_part = denom_part.replace(duration, 'd')
            break
    
    if 'm-2' in denom_part: # area units
        prefix, _ = extract_prefix(denom_part, 'area')
        if prefix != 'c':   # convert to cm-2
            rate_val /= (PREFIXES.get(prefix, 1.0) / PREFIXES['c']) ** 2
            denom_part = denom_part.replace(f"{prefix}m-2", "cm-2")
    
    elif 'g' in denom_part and not (denom_part.startswith('d') and len(denom_part) <= 3):  # mass units, avoiding 'day' confusion
        prefix, _ = extract_prefix(denom_part, 'mass')
        rate_val /= PREFIXES.get(prefix, 1.0)
        denom_part = denom_part.replace(f"{prefix}g", "g")
    
    elif re.search(r'[nuμmcdk]{2}-2', denom_part):  # duplicate character units e.g. mm
        prefix = denom_part[0]
        rate_val /= (PREFIXES.get(prefix, 1.0) / PREFIXES['c']) ** 2    # convert to cm-2
        denom_part = "cm-2" + denom_part[3:]
    
    return rate_val, denom_part

def rate_conversion(
    rate_val: float, 
    rate_error: float = None,
    rate_unit: str = "", 
) -> [tuple[float, str], tuple[float, str, float]]:
    """
    Convert rate to standardized units (gCaCO3 per day) and propagate errors if provided.
    
    Parameters:
    - rate_val: Rate value to convert
    - rate_unit: Original rate unit string (e.g., 'mgCaCO3 cm-2d-1')
    - rate_error: Standard deviation or standard error of the rate (optional)
    
    Returns:
    - Converted rate value
    - New standardized rate unit
    - Converted error value (if rate_error was provided)
    """
    if rate_unit is None or rate_unit != rate_unit: # handle nans
        if rate_error is not None:
            return rate_val, "", rate_error
        return rate_val, ""
    
    try:    # split into numerator and denominator
        num_part, denom_part = parse_unit_components(rate_unit)
    except ValueError as e:
        if rate_error is not None:
            return rate_val, str(e), rate_error
        return rate_val, str(e)
    
    original_val = rate_val    
    rate_val, new_num = convert_numerator(num_part, rate_val)    
    rate_val, new_denom = convert_denominator(denom_part, rate_val)
    new_rate_unit = f"{new_num} {new_denom}"
    
    # calculate the scaling factor and propagate error if provided
    if rate_error is not None:
        if original_val != 0:
            scaling_factor = rate_val / original_val
            new_error = rate_error * abs(scaling_factor)
        else:
            # if original value is zero, can't determine scaling factor
            # In this case, we assume the error scales similarly
            new_error = rate_error
        return rate_val, new_error, new_rate_unit
    
    return rate_val, None, new_rate_unit


### sensitivity analysis
def select_by_stat(ds, variables_stats: dict):
    """
    Selects values from an xarray dataset based on the specified statistics for given variables.
    
    Parameters:
        ds (xr.Dataset): The input dataset with a 'param_combination' dimension and coordinates for the variables.
        variables_stats (dict): A dictionary where keys are variable names and values are the statistics to use for selection ('min', 'max', 'mean').
    
    Returns:
        xr.Dataset: Dataset subset at the selected coordinate values.
    """
    selected_coords = {}
    
    for var, stat in variables_stats.items():
        if stat == "min":
            selected_coords[var] = ds[var].min().item()
        elif stat == "max":
            selected_coords[var] = ds[var].max().item()
        elif stat == "mean":
            selected_coords[var] = ds[var].mean().item()
        else:
            raise ValueError(f"stat for {var} must be 'min', 'max', or 'mean'")
    
    # Select the closest matching values
    ds_selected = ds.sel(selected_coords, method="nearest")
    
    return ds_selected





### DEPRECATED

# Function to determine optimal number of clusters using silhouette score
# def optimal_kmeans(data, max_clusters=8):
#     best_k = 2  # Minimum sensible number of clusters
#     best_score = -1
#     scores = []

#     for k in range(2, min(len(data), max_clusters + 1)):  # Avoid excessive clustering
#         kmeans = KMeans(n_clusters=k, random_state=42, n_init=max_clusters)
#         labels = kmeans.fit_predict(data)
#         score = silhouette_score(data, labels)
#         scores.append((k, score))

#         if score > best_score:
#             best_score = score
#             best_k = k

#     return best_k, scores


# def rate_conversion(rate_val: float, rate_unit: str) -> tuple[float, str]:
#     """Conversion into gCaCO3 ... day-1 for absolute rates"""
#     rate_components = rate_unit.split(' ')
#     # print(rate_components)
#     num = rate_components[0]
#     denom = rate_components[1]
    
#     # if rate_unit == 'um d-1':
#     #     print('problem')
#     # process numerator
#     if 'delta' in num:  # no conversion needed
#         # rate_val = rate_val
#         new_rate_unit_num = num.replace('delta%', 'delta%')
#     else:
#         # Initialize num_prefix as an empty string
#         num_prefix = ""
        
#         new_rate_unit_num = num
#         if 'mol' in num:    # moles
#             rate_val *= MOLAR_MASS_CACO3
#             num_prefix = num.split('mol')[0]
#             num = num.replace('mol', 'g')
#             new_rate_unit_num = num.replace(num_prefix, '')
    
#         elif 'g' in num:    # mass
#             num_prefix = num.split('g')[0]
#             # new_rate_unit_num = num.replace(num_prefix, 'g')
#             new_rate_unit_num = num.replace(num_prefix, '')
#         elif 'm2' in num:    # area (need to square)
#             num_prefix = num.split('m2')[0]
#             # if more than one 'num_prefix' occurence, do this
#             if num.count(num_prefix) > 1 and num_prefix != '':
#                 new_rate_unit_num = num[1:]
#         elif 'm' in num:    # extension
#             num_prefix = num[0] # TODO: always valid?
#             if num.count(num_prefix) > 1:   # mm
#                 new_rate_unit_num = num[1:]
#             else:
#                 new_rate_unit_num = num.replace(num_prefix, '')
                    
#         # check if prefix exists and is valid
#         if num_prefix in PREFIXES:
#             if 'm2' in num:  # area
#                 rate_val *= PREFIXES[num_prefix]**2
#             else:
#                 rate_val *= PREFIXES[num_prefix]
        
#     denom_prefix = ""
#     new_rate_unit_denom = denom
#     # process denominator
#     if 'm-2' in denom:  # specific area
#         denom_prefix = denom.split('m-2')[0]
#         if denom.count(denom_prefix) > 1 and denom_prefix != '':    # mm
#             new_rate_unit_denom = 'c'+denom[1:]
#             denom_prefix = 'm'
#         else:
#             new_rate_unit_denom = denom.replace(denom_prefix, 'c')
                
#         if denom_prefix in PREFIXES:
#             rate_val /= (PREFIXES[denom_prefix]/PREFIXES['c'])**2
#     elif 'g' in denom:  # specific mass
#         denom_prefix = denom.split('g')[0]
#         if denom_prefix in PREFIXES:
#             rate_val /= PREFIXES[denom_prefix]

#         new_rate_unit_denom = denom.replace(denom_prefix, '')
#     # time conversion
#     if any(duration in denom for duration in DURATIONS.keys()):
#         for duration, factor in DURATIONS.items():
#             if duration in denom:
#                 rate_val *= factor
#                 new_rate_unit_denom = new_rate_unit_denom.replace(duration, 'd')
#                 break
            
#     new_rate_unit = f"{new_rate_unit_num} {new_rate_unit_denom}"

#     return rate_val, new_rate_unit   

    
    
    
    # def uniquify_dois(df):
#     """
#     Uniquify dois to reflect locations (for studies with multiple locations)
    
#     Args:
#         df (pd.DataFrame): dataframe containing doi and location columns
    
#     Returns:
#         pd.DataFrame: dataframe with uniquified dois, and copies of original
#     """
#     df['original_doi'] = df['doi']
#     temp_df = df.copy()
#     temp_df['location_lower'] = temp_df['location'].str.lower()
    
    
    
#     unique_locs = temp_df.drop_duplicates(['location_lower', 'coords', 'cleaned_coords'])[['doi', 'location']]
#     # unique_locs.dropna(subset=['latitude', 'longitude'], inplace=True) # remove empty rows
#     dois = unique_locs.doi
#     temp_df.index = uniquify_repeated_values(dois)
#     doi_location_map = dict(zip(zip(temp_df.drop_duplicates(subset=['doi', 'location_lower', 'coords', 'cleaned_coords'])['doi'], 
#                                 temp_df.drop_duplicates(subset=['doi', 'location_lower', 'coords', 'cleaned_coords'])['location']), 
#                             dois))
#     temp_df['doi'] = [doi_location_map.get((doi, loc), doi) for doi, loc in zip(temp_df['doi'], temp_df['location'])]
#     # temp_df['doi'] = temp_df.index
#     return temp_df

    
    
    # unique_locs.doi = utils.uniquify_repeated_values(df.drop_duplicates(subset=['doi', 'location']).doi)

    # # create a dictionary mapping from original (doi, location) pairs to uniquified dois
    # doi_location_map = dict(zip(zip(df.drop_duplicates(subset=['doi', 'location'])['doi'], 
    #                             df.drop_duplicates(subset=['doi', 'location'])['location']), 
    #                         unique_locs['doi']))
    # df['doi'] = [doi_location_map.get((doi, loc), doi) for doi, loc in zip(df['doi'], df['location'])]
    
    # return df