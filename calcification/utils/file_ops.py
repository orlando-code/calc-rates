# file handling
import datetime
import re

# general
import numpy as np
import pandas as pd

# R
import rpy2
import rpy2.robjects.packages as rpackages
import yaml
from openpyxl import load_workbook


def _convert_numpy(obj) -> dict | list | np.ndarray | float | int | str:
    """Convert numpy types to native Python types for safe YAML serialization.

    Args:
        obj: Object that might contain numpy data types

    Returns:
        Object with numpy types converted to Python native types
    """
    match obj:
        case dict():
            return {k: _convert_numpy(v) for k, v in obj.items()}
        case list() | tuple():
            return [_convert_numpy(i) for i in obj]
        case np.ndarray():
            return obj.tolist()
        case np.number():
            return obj.item()
        case np.float32() | np.float64():
            return float(obj)
        case np.int32() | np.int64():
            return int(obj)
        case _:
            return obj


def read_yaml(yaml_fp) -> dict:
    """Read in yaml file"""
    with open(yaml_fp, "r") as stream:
        try:
            return yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)
            return None


def write_yaml(data: dict, fp="unnamed.yaml") -> None:
    """Safe writing to yaml files"""
    # convert numpy values to Python native types before serialization
    converted_data = _convert_numpy(data)
    with open(fp, "w") as file:
        yaml.dump(converted_data, file)


def append_to_yaml(data: dict, fp="unnamed.yaml") -> None:
    """Append dictionary to yaml file"""
    converted_data = _convert_numpy(data)
    with open(fp, "a") as file:
        yaml.dump(converted_data, file)


def ensure_r_package_imported(package_name: str) -> None:
    """Ensure that an R package is imported using rpy2"""
    try:
        rpackages.importr(package_name)
    except rpy2.robjects.packages.PackageNotFoundError:
        rpackages.importr(package_name)


def get_now_timestamp_formatted():
    """Get a nicely-formatted timestamp for file naming."""
    return datetime.datetime.now().strftime("%Y-%m-%d--%H-%M-%S")


def extract_year_from_str(s) -> pd.Timestamp:
    """Extract year from a string containing a year (e.g. 2022). Useful for cases where 'authors' field
    has been provided with year attached to it.

    Args:
        s (str): string containing year.

    Returns:
        pd.Timestamp: year extracted from string.

    Raises:
        ValueError: If multiple years are found in the string.
    """
    try:
        matches = re.findall(r"\d{4}", str(s))
        if not matches:
            return pd.NaT
        if len(matches) > 1:
            raise ValueError(f"Multiple years found in string: {matches} (input: {s})")
        return pd.to_datetime(matches[0], format="%Y")
    except ValueError as e:
        raise e


def get_highlighted(
    fp: str, sheet_name: str, rgb_color: str = "FFFFC000"
) -> pd.DataFrame:
    """
    Loads an Excel sheet into a DataFrame and masks non-highlighted values with NaN,
    except for a specified subset of columns.

    Args:
        fp (str): path to the Excel file.
        sheet_name (str): name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).

    Returns:
        pd.DataFrame: DataFrame with non-highlighted values masked as NaN.
    """
    df = pd.read_excel(fp, sheet_name=sheet_name, engine="openpyxl", dtype=str)
    mask = get_highlighted_mask(
        fp, sheet_name, rgb_color
    )  # generate mask of highlighted cells
    return df.where(
        mask.drop(0).reset_index(drop=True), np.nan
    )  # remove first row of mask and reset index


def get_highlighted_mask(
    fp: str, sheet_name: str, rgb_color: str = "FFFFC000"
) -> pd.DataFrame:
    """
    Creates a boolean mask DataFrame where True indicates a highlighted cell.

    Args:
        fp (str): path to the Excel file.
        sheet_name (str): name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).

    Returns:
        pd.DataFrame: boolean mask DataFrame with True for highlighted cells.
    """
    wb = load_workbook(fp, data_only=True, read_only=True)
    ws = wb[sheet_name]

    df = pd.DataFrame([[cell.value for cell in row] for row in ws.iter_rows()])
    df = df.dropna(axis=1, how="all")  # remove any empty columns
    df.columns = df.iloc[0]  # set the first row as the column headers
    # df = df.drop(0)  # remove the header row  # TODO: check this functionality
    mask = pd.DataFrame(
        False, index=df.index, columns=df.columns
    )  # mask with False by default, same shape as df
    # mask.drop(0, inplace=True)  # drop the first row (header)

    # set highlighted cells to True, all else False
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == rgb_color:
                mask.iat[row_idx, col_idx] = True

    # mark 'include' and 'n' columns as True for future processing
    for col in ["Include", "n", "Species types", "Location"]:
        if col in mask.columns:
            mask[col] = True

    return mask
