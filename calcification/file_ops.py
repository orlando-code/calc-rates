# file handling
import yaml
import datetime
from openpyxl import load_workbook

# general
import numpy as np
import pandas as pd

# R
import rpy2


def read_yaml(yaml_fp) -> dict:
    """Read in yamls"""
    with open(yaml_fp, "r") as stream:
        try:
            return yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)
            return None
        
        
def _convert_numpy(obj):
    """Convert numpy types to native Python types for YAML serialization.
    
    Args:
        obj: Object that might contain numpy data types
        
    Returns:
        Object with numpy types converted to Python native types
    """
    if isinstance(obj, dict):
        return {k: _convert_numpy(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_convert_numpy(i) for i in obj]
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, np.number):
        return obj.item()
    elif isinstance(obj, (np.float32, np.float64)):
        return float(obj)
    elif isinstance(obj, (np.int32, np.int64)):
        return int(obj)
    else:
        return obj


def write_yaml(data: dict, fp="unnamed.yaml") -> None:
    """Safe writing to yaml files"""
    # Convert numpy values to Python native types before serialization
    converted_data = _convert_numpy(data)
    with open(fp, "w") as file:
        yaml.dump(converted_data, file)


def append_to_yaml(data: dict, fp="unnamed.yaml") -> None:
    """Append to yaml"""
    converted_data = _convert_numpy(data)
    with open(fp, "a") as file:
        yaml.dump(converted_data, file)


def ensure_r_package_imported(package_name):
    """
    Ensure that an R package is imported using rpy2.
    
    Parameters:
    -----------
    package_name : str
        Name of the R package to import.
    
    Returns:
    --------
    None
    """
    try:
        rpackages.importr(package_name)
    except rpy2.robjects.packages.PackageNotFoundError:
        rpackages.importr(package_name)
        

def get_formatted_timestamp():
    """Get a nicely-formatted timestamp for file naming."""
    return datetime.datetime.now().strftime("%Y-%m-%d--%H-%M-%S")


def extract_year_from_str(s) -> pd.Timestamp:
    """Extract year from a string. Useful for cases where 'authors' has been provided with attached year.
    
    Args:
        s (str): String containing year.
        
    Returns:
        pd.Timestamp: Year extracted from string.
    """
    try:
        match = re.search(r'\d{4}', str(s))
        if match:
            return pd.to_datetime(match.group(), format='%Y')
        return pd.NaT
    except ValueError:
        return None
    
    
def get_highlighted(fp: str, sheet_name: str, rgb_color: str = "FFFFC000") -> pd.DataFrame:
    """
    Loads an Excel sheet into a DataFrame and masks non-highlighted values with NaN,
    except for a specified subset of columns.

    Args:
        fp (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).
        keep_cols (list): List of column names to keep unchanged.

    Returns:
        pd.DataFrame: DataFrame with non-highlighted values masked as NaN.
    """
    df = pd.read_excel(fp, sheet_name=sheet_name, engine="openpyxl", dtype=str)
    mask = get_highlighted_mask(fp, sheet_name, rgb_color)  # generate mask of highlighted cells
    return df.where(mask.drop(0).reset_index(drop=True), np.nan)     # remove first row of mask and reset index


def get_highlighted_mask(fp: str, sheet_name: str, rgb_color: str = "FFFFC000") -> pd.DataFrame:
    """
    Creates a boolean mask DataFrame where True indicates a highlighted cell.

    Args:
        fp (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet.
        rgb_color (str): RGB color to filter (e.g., 'FFFFC000' for calculated values).

    Returns:
        pd.DataFrame: Boolean mask DataFrame with True for highlighted cells.
    """
    wb = load_workbook(fp, data_only=True, read_only=True)
    ws = wb[sheet_name]

    df = pd.DataFrame([[cell.value for cell in row] for row in ws.iter_rows()])
    df = df.dropna(axis=1, how='all')   # remove any empty columns
    df.columns = df.iloc[0]  # set the first row as the column headers
    # df = df.drop(0)  # remove the header row
    mask = pd.DataFrame(False, index=df.index, columns=df.columns)  # mask with False by default, same shape as df
    # mask.drop(0, inplace=True)  # drop the first row (header)

    # set highlighted cells to True, all else False
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb == rgb_color:
                mask.iat[row_idx, col_idx] = True
                
    # Always mark 'include' and 'n' columns as True for future processing
    for col in ['Include', 'n', 'Species types', 'Location']:
        if col in mask.columns:
            mask[col] = True

    return mask